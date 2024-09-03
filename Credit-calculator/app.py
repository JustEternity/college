import datetime
import sys
import openpyxl



from Credit_calculator import Ui_MainWindow
from info_window import Ui_Info
from datetime import datetime
from dateutil.relativedelta import relativedelta
from PyQt6.QtWidgets import QApplication, QMainWindow, QButtonGroup, QMessageBox, QTableWidgetItem, QFileDialog
from payment_schedule import Ui_Schedule
from PyQt6.QtCore import Qt
from decimal import Decimal, getcontext, ROUND_DOWN





class Calculator_app(QMainWindow, Ui_MainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(self)

        self.sum = self.enter_sum.text()
        self.rate = self.rate_enter.text()
        self.term = self.enter_term.text()
        self.date = self.dateEdit.date()

        self.enter_sum.textChanged.connect(self.change_sum)
        self.rate_enter.textChanged.connect(self.change_rate)
        self.enter_term.textChanged.connect(self.change_term)
        self.dateEdit.dateChanged.connect(self.change_date)

        self.radioButton_group = QButtonGroup()
        self.radioButton_group.addButton(self.radioButton_1, id=1)
        self.radioButton_group.addButton(self.radioButton_2, id=2)

        self.calculation_button.clicked.connect(self.calc_button_click)
        self.info_button.clicked.connect(self.show_info)
        self.paygraph_button.clicked.connect(self.calc_schedule_and_res)


    def calc_schedule_and_res(self):
        self.calc_button_click()
        self.draw_schedule()

    def calc_button_click(self):
        if self.sum and self.date and self.rate and self.term:
            match self.radioButton_group.checkedId():
                case 1:
                    self.calculate_annuty()
                case 2:
                    self.calculate_diff()
                case -1:
                    QMessageBox.warning(self.layoutWidget, "Предупреждение", "Выберите тип платежа по кредиту")
        else:
            QMessageBox.warning(self.layoutWidget, "Предупреждение", "Заполните поля выше")

    def change_sum(self, text):
        if text:
            try:
                self.sum = int(text)
                if self.sum > 100000000:
                    self.sum = 100000000
                    self.enter_sum.setText('100000000')
            except ValueError:
                QMessageBox.warning(self.layoutWidget, "Предупреждение", "Здесь должно быть целое число")
        else:
            self.sum = 0

    def change_rate(self, text):
        if text:
            try:
                self.rate = float(text.replace(',', '.'))
                if self.rate > 292:
                    self.rate = 292
                    self.rate_enter.setText('292')
            except ValueError:
                QMessageBox.warning(self.layoutWidget, "Предупреждение", "Здесь должно быть число")
        else:
            self.rate = 0

    def change_term(self, text):
        if text:
            try:
                self.term = int(text)
                if self.choice_term.currentText() == 'мес.' and self.term > 360:
                    self.term = 360
                    self.enter_term.setText('360')
                if self.choice_term.currentText() == 'лет' and self.term > 30:
                    self.term = 30
                    self.enter_term.setText('30')
            except ValueError:
                QMessageBox.warning(self.layoutWidget, "Предупреждение", "Здесь должно быть число")
        else:
            self.term = 0

    def change_date(self, text):
        self.date = text

    def calculate_annuty(self):
        ''' Функция, рассчитывающая ежемесячный платеж, общую сумму кредита и сумму процентов для аннуитентного кредита
        '''
        self.type_term = 1 if self.choice_term.currentText() == 'мес.' else 12
        self.month_rate = self.rate / 1200
        self.month_payment_annuty = self.sum*((self.month_rate * (1 + self.month_rate)**(self.term*self.type_term)) /
                                              ((1 + self.month_rate)**(self.term*self.type_term)  - 1))
        self.month_payment.setText(f"Ежемесячный платеж: {self.month_payment_annuty:.2f}")
        self.schedule_ann = self.calculate_annuity_payments(self.sum, self.month_payment_annuty, self.rate/100,
                                              self.term*self.type_term, self.date.toString('yyyy-MM-dd'))
        self.percents_annuity = self.percents_sum(self.schedule_ann)
        self.overpayment.setText(f'Сумма переплаты: {self.percents_annuity:.2f}')
        self.amount_payments.setText(f'Сумма выплат: {self.percents_annuity + self.sum:.2f}')

    def calculate_annuity_payments(self, loan_amount, monthly_payment, interest_rate, loan_term, issue_date):
        getcontext().prec = 28
        payment_schedule = {}
        remaining_loan = Decimal(loan_amount)
        monthly_payment = Decimal(monthly_payment)
        interest_rate = Decimal(interest_rate)
        issue_date = datetime.strptime(issue_date, "%Y-%m-%d")

        for month in range(1, loan_term + 1):
            payment_date = issue_date + relativedelta(months=+month)
            if payment_date.day < issue_date.day:
                payment_date = payment_date + relativedelta(day=31)
            interest_payment = remaining_loan * (interest_rate / 12)
            principal_payment = monthly_payment - interest_payment
            remaining_loan -= principal_payment
            if remaining_loan < Decimal("0.0000000001"):
                remaining_loan = Decimal("0.00")
            payment_schedule[payment_date.strftime("%Y-%m-%d")] = [f'{(monthly_payment.quantize(Decimal("0.00"), rounding=ROUND_DOWN))}',
                                                               f'{(interest_payment.quantize(Decimal("0.00"), rounding=ROUND_DOWN))}',
                                                               f'{(remaining_loan.quantize(Decimal("0.00"), rounding=ROUND_DOWN))}']
            if payment_schedule[payment_date.strftime("%Y-%m-%d")][-1] == '0.00':
                break

        self.income.setText(f'Необходимый уровень дохода для одобрения кредита:\n{monthly_payment*2:.0f}')

        return payment_schedule

    def calculate_diff(self):
        ''' Функция для расчета суммы основного долга для каждого платежа, ежемесячный платеж уменьшается со временем,
        поэтому он показывается только в разделе "График платежей"
        '''
        self.type_term = 1 if self.choice_term.currentText() == 'мес.' else 12
        self.start_date = datetime(self.date.year(), self.date.month(), self.date.day())
        if self.type_term == 12:
            self.end_date = self.start_date + relativedelta(years=self.term)
        else:
            self.end_date = self.start_date + relativedelta(months=self.term)

        self.count_days = (self.end_date - self.start_date).days

        # Часть основного долга, одинаковая для каждого месяца (платежа)
        self.part_of_debt = self.sum / (self.term*self.type_term)


        self.amount_payments.setText(f"Ежемесячный платеж\nбез учета процентов: {self.part_of_debt:.2f}")

        self.schedule_differ = self.schedule_diff(self.sum, self.rate, self.term*self.type_term, self.date)
        self.sum_of_percents = self.percents_sum(self.schedule_differ)
        self.month_payment.setText(f"Переплата: {self.sum_of_percents:.2f}")
        self.overpayment.setText(f"Сумма выплат: {self.sum_of_percents + self.sum:.2f}")

    def percents_sum(self, schedule):
        return sum(float(value[1]) for value in schedule.values())

    def schedule_diff(self, loan_amount, interest_rate, term, issue_date):
        payments = {}
        monthly_interest_rate = interest_rate / 12 / 100
        remaining_balance = loan_amount

        # Преобразование QDate в datetime
        issue_date = datetime(issue_date.year(), issue_date.month(), issue_date.day())

        for month in range(1, term + 1):
            interest_payment = remaining_balance * monthly_interest_rate
            principal_payment = loan_amount / term
            total_payment = principal_payment + interest_payment
            remaining_balance -= principal_payment

            # Adjust the payment date if necessary
            payment_date = (issue_date + relativedelta(months=+month)).date()
            if payment_date.day != issue_date.day:
                payment_date = payment_date.replace(day=min(issue_date.day, payment_date.day))

            # Store the payment details in the dictionary
            payments[payment_date.strftime("%Y-%m-%d")] = [f'{abs(round(total_payment, 2))}',
                                                           f'{abs(round(interest_payment, 2))}',
                                                           f'{abs(round(remaining_balance, 2))}']

        self.income.setText(f'Необходимый уровень дохода для одобрения кредита:\n {float(next(iter(payments.values()))[0])*2:.0f}')

        return payments

    def draw_schedule(self):
        if self.sum and self.date and self.rate and self.term:
            match self.radioButton_group.checkedId():
                case 1:
                    self.calculate_paygraph(self.schedule_ann)
                case 2:
                    self.calculate_paygraph(self.schedule_differ)


    def calculate_paygraph(self, schedule):
        ''' Функция для вывода графика платежей по кредиту
        '''
        self.graph_widget = Graph_window()
        self.graph_widget.show()
        self.graph_widget.display_schedule(schedule)

    def show_info(self):
        ''' Функция для вывода справочной информации по кредитованию
        '''
        self.info_widget = Info_window()
        self.info_widget.show()



class Info_window(QMainWindow, Ui_Info):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(self)



class Graph_window(QMainWindow, Ui_Schedule):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(self)
        self.save_button.clicked.connect(self.save_xls)

    def display_schedule(self, data_dict):
        try:
            row = 0
            self.tableWidget.setRowCount(len(data_dict))  # Установка количества строк после цикла
            row = 0
            for key, values in data_dict.items():
                date = QTableWidgetItem(key)
                date.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tableWidget.setItem(row, 0, date)  # Установка ключа в первый столбец
                for col, value in enumerate(values):
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.tableWidget.setItem(row, col + 1, item)  # Установка значений в следующие столбцы
                row += 1
        except ValueError:
            QMessageBox.warning(self.layout, 'Warning', 'Enter data to procceding')


    def save_xls(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_name:
            wb = openpyxl.Workbook()
            ws = wb.active

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            # Вставить 1 строку перед экспортируемой таблицей
            ws.insert_rows(1, 1)

            # Записать слова в ячейки
            ws.cell(row=1, column=1).value = "Дата"
            ws.cell(row=1, column=2).value = "Сумма платежа"
            ws.cell(row=1, column=3).value = "Проценты"
            ws.cell(row=1, column=4).value = "Остаток"

            # Записать данные таблицы
            for row in range(self.tableWidget.rowCount()):
                for column in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, column)
                    if item is not None:
                        ws.cell(row=row+2, column=column+1, value=item.text())

            # Сохранить файл
            wb.save(file_name)


def main():
    app = QApplication(sys.argv)
    window = Calculator_app()
    window.show()
    app.exec()

if __name__ == "__main__":
    main()